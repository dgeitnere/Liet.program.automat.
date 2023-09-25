from openpyxl import Workbook, load_workbook 
wb=load_workbook('./1.darbs/tests/test1.xlsx')
ws=wb.active
total=0

max_row = ws.max_row
for row in range (2,max_row+1):
    rate=ws['B'+str(row)].value
    hours=ws['C'+str(row)].value
    id=ws['A'+str(row)].value
    if (type(rate) != str and type (hours) != str):
        salary = float(hours) * float(rate)
        if salary > 3000:
         total = (total + 1)
print(total)
wb.close()
